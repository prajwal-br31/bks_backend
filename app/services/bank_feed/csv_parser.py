"""Server-side CSV and Excel parsers for bank statement imports."""

import csv
import io
import logging
import re
import zipfile
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, List, Dict, Any, Union

logger = logging.getLogger(__name__)

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logger.warning("openpyxl not installed, Excel support disabled")


@dataclass
class ParsedTransaction:
    """Represents a parsed bank transaction."""
    date: datetime
    description: str
    amount: float
    type: str  # "credit" or "debit"
    balance: Optional[float] = None
    category: Optional[str] = None
    check_number: Optional[str] = None
    post_date: Optional[datetime] = None
    external_id: Optional[str] = None
    memo: Optional[str] = None
    raw_data: Dict[str, Any] = field(default_factory=dict)
    row_number: int = 0


@dataclass
class ParseResult:
    """Result of parsing a bank statement file."""
    transactions: List[ParsedTransaction]
    total_rows: int
    parsed_rows: int
    skipped_rows: int
    errors: List[str]
    warnings: List[str]
    bank_name: Optional[str] = None
    account_last4: Optional[str] = None
    statement_start: Optional[datetime] = None
    statement_end: Optional[datetime] = None


class BankCsvParser(ABC):
    """Abstract base class for bank CSV parsers."""

    @abstractmethod
    def parse(self, content: bytes, filename: str = "") -> ParseResult:
        """Parse CSV content and return transactions."""
        pass

    @abstractmethod
    def detect(self, content: bytes) -> bool:
        """Detect if this parser can handle the given content."""
        pass

    def _decode_content(self, content: bytes) -> str:
        """Decode bytes to string with multiple encoding attempts."""
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252", "iso-8859-1"]
        for encoding in encodings:
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        return content.decode("utf-8", errors="replace")

    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """Parse date string with multiple format attempts."""
        if not date_str:
            return None
        
        date_str = date_str.strip()
        formats = [
            "%m/%d/%Y",
            "%m/%d/%y",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m-%d-%Y",
            "%Y/%m/%d",
            "%b %d, %Y",
            "%B %d, %Y",
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        logger.warning(f"Could not parse date: {date_str}")
        return None

    def _parse_amount(self, amount_str: str) -> Optional[float]:
        """Parse amount string to float."""
        if not amount_str:
            return None
        
        # Remove currency symbols and whitespace
        cleaned = re.sub(r"[^\d.\-\(\)]", "", amount_str.strip())
        
        # Handle parentheses as negative
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        
        try:
            return float(cleaned)
        except ValueError:
            logger.warning(f"Could not parse amount: {amount_str}")
            return None


class ChaseCsvParser(BankCsvParser):
    """Parser for Chase Bank CSV exports."""

    CHASE_COLUMNS = [
        "Details", "Posting Date", "Description", "Amount", "Type", "Balance", "Check or Slip #"
    ]

    def detect(self, content: bytes) -> bool:
        """Detect if content is a Chase CSV."""
        try:
            text = self._decode_content(content)
            first_line = text.split("\n")[0].strip()
            
            # Check for Chase column headers
            if any(col in first_line for col in ["Posting Date", "Details", "Check or Slip"]):
                return True
            
            return False
        except Exception:
            return False

    def parse(self, content: bytes, filename: str = "") -> ParseResult:
        """Parse Chase CSV content."""
        transactions = []
        errors = []
        warnings = []
        skipped_rows = 0
        
        try:
            text = self._decode_content(content)
            reader = csv.DictReader(io.StringIO(text))
            
            row_num = 0
            min_date = None
            max_date = None
            
            for row in reader:
                row_num += 1
                
                try:
                    # Parse posting date
                    date_str = row.get("Posting Date", "") or row.get("Date", "")
                    date = self._parse_date(date_str)
                    
                    if not date:
                        warnings.append(f"Row {row_num}: Could not parse date '{date_str}'")
                        skipped_rows += 1
                        continue
                    
                    # Track date range
                    if min_date is None or date < min_date:
                        min_date = date
                    if max_date is None or date > max_date:
                        max_date = date
                    
                    # Parse description
                    description = row.get("Description", "").strip()
                    if not description:
                        warnings.append(f"Row {row_num}: Empty description")
                        description = "Unknown Transaction"
                    
                    # Parse amount
                    amount_str = row.get("Amount", "")
                    amount = self._parse_amount(amount_str)
                    
                    if amount is None:
                        warnings.append(f"Row {row_num}: Could not parse amount '{amount_str}'")
                        skipped_rows += 1
                        continue
                    
                    # Determine type
                    type_field = row.get("Type", "").lower()
                    if type_field in ("credit", "deposit", "cr"):
                        txn_type = "credit"
                    elif type_field in ("debit", "withdrawal", "dr"):
                        txn_type = "debit"
                    else:
                        # Infer from amount
                        txn_type = "credit" if amount > 0 else "debit"
                    
                    # Normalize amount to positive
                    amount = abs(amount)
                    
                    # Parse balance
                    balance_str = row.get("Balance", "")
                    balance = self._parse_amount(balance_str)
                    
                    # Parse check number
                    check_number = row.get("Check or Slip #", "").strip() or None
                    
                    # Create transaction
                    txn = ParsedTransaction(
                        date=date,
                        description=description,
                        amount=amount,
                        type=txn_type,
                        balance=balance,
                        check_number=check_number,
                        raw_data=dict(row),
                        row_number=row_num,
                    )
                    
                    transactions.append(txn)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    skipped_rows += 1
            
            return ParseResult(
                transactions=transactions,
                total_rows=row_num,
                parsed_rows=len(transactions),
                skipped_rows=skipped_rows,
                errors=errors,
                warnings=warnings,
                bank_name="Chase",
                statement_start=min_date,
                statement_end=max_date,
            )
            
        except Exception as e:
            logger.error(f"Error parsing Chase CSV: {e}")
            return ParseResult(
                transactions=[],
                total_rows=0,
                parsed_rows=0,
                skipped_rows=0,
                errors=[f"Failed to parse file: {str(e)}"],
                warnings=[],
            )


class GenericCsvParser(BankCsvParser):
    """Generic CSV parser that attempts to detect column mappings."""

    # Common column name variations
    DATE_COLUMNS = ["date", "trans date", "transaction date", "posting date", "post date", "value date"]
    DESC_COLUMNS = ["description", "desc", "memo", "narrative", "particulars", "details", "payee"]
    AMOUNT_COLUMNS = ["amount", "amt", "value", "sum", "transaction amount"]
    CREDIT_COLUMNS = ["credit", "deposit", "cr", "credits", "money in"]
    DEBIT_COLUMNS = ["debit", "withdrawal", "dr", "debits", "money out"]
    BALANCE_COLUMNS = ["balance", "running balance", "available balance", "ledger balance"]
    TYPE_COLUMNS = ["type", "transaction type", "dr/cr", "debit/credit"]

    def detect(self, content: bytes) -> bool:
        """Generic parser can handle any CSV as fallback."""
        try:
            text = self._decode_content(content)
            # Check if it looks like CSV
            reader = csv.reader(io.StringIO(text))
            first_row = next(reader, None)
            return first_row is not None and len(first_row) >= 2
        except Exception:
            return False

    def _find_column(self, headers: List[str], candidates: List[str]) -> Optional[str]:
        """Find matching column from candidates."""
        headers_lower = {h.lower().strip(): h for h in headers}
        for candidate in candidates:
            if candidate in headers_lower:
                return headers_lower[candidate]
        return None

    def parse(self, content: bytes, filename: str = "") -> ParseResult:
        """Parse generic CSV content."""
        transactions = []
        errors = []
        warnings = []
        skipped_rows = 0
        
        try:
            text = self._decode_content(content)
            reader = csv.DictReader(io.StringIO(text))
            headers = reader.fieldnames or []
            
            # Map columns
            date_col = self._find_column(headers, self.DATE_COLUMNS)
            desc_col = self._find_column(headers, self.DESC_COLUMNS)
            amount_col = self._find_column(headers, self.AMOUNT_COLUMNS)
            credit_col = self._find_column(headers, self.CREDIT_COLUMNS)
            debit_col = self._find_column(headers, self.DEBIT_COLUMNS)
            balance_col = self._find_column(headers, self.BALANCE_COLUMNS)
            type_col = self._find_column(headers, self.TYPE_COLUMNS)
            
            if not date_col:
                return ParseResult(
                    transactions=[],
                    total_rows=0,
                    parsed_rows=0,
                    skipped_rows=0,
                    errors=["Could not identify date column"],
                    warnings=[],
                )
            
            if not desc_col:
                warnings.append("Could not identify description column, using first text column")
                # Find first text-like column
                for h in headers:
                    if h != date_col and h not in [amount_col, credit_col, debit_col, balance_col]:
                        desc_col = h
                        break
            
            row_num = 0
            min_date = None
            max_date = None
            
            for row in reader:
                row_num += 1
                
                try:
                    # Parse date
                    date_str = row.get(date_col, "") if date_col else ""
                    date = self._parse_date(date_str)
                    
                    if not date:
                        skipped_rows += 1
                        continue
                    
                    if min_date is None or date < min_date:
                        min_date = date
                    if max_date is None or date > max_date:
                        max_date = date
                    
                    # Parse description
                    description = row.get(desc_col, "Unknown") if desc_col else "Unknown"
                    
                    # Parse amount and determine type
                    amount = None
                    txn_type = "debit"
                    
                    if amount_col:
                        amount = self._parse_amount(row.get(amount_col, ""))
                        if amount is not None:
                            txn_type = "credit" if amount > 0 else "debit"
                            amount = abs(amount)
                    elif credit_col and debit_col:
                        credit = self._parse_amount(row.get(credit_col, ""))
                        debit = self._parse_amount(row.get(debit_col, ""))
                        
                        if credit and credit > 0:
                            amount = credit
                            txn_type = "credit"
                        elif debit and debit > 0:
                            amount = debit
                            txn_type = "debit"
                    
                    if amount is None:
                        skipped_rows += 1
                        continue
                    
                    # Override type if column exists
                    if type_col and row.get(type_col):
                        type_val = row.get(type_col, "").lower()
                        if "credit" in type_val or "cr" in type_val or "deposit" in type_val:
                            txn_type = "credit"
                        elif "debit" in type_val or "dr" in type_val or "withdrawal" in type_val:
                            txn_type = "debit"
                    
                    # Parse balance
                    balance = self._parse_amount(row.get(balance_col, "")) if balance_col else None
                    
                    txn = ParsedTransaction(
                        date=date,
                        description=description.strip(),
                        amount=amount,
                        type=txn_type,
                        balance=balance,
                        raw_data=dict(row),
                        row_number=row_num,
                    )
                    
                    transactions.append(txn)
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    skipped_rows += 1
            
            return ParseResult(
                transactions=transactions,
                total_rows=row_num,
                parsed_rows=len(transactions),
                skipped_rows=skipped_rows,
                errors=errors,
                warnings=warnings,
                bank_name="Unknown",
                statement_start=min_date,
                statement_end=max_date,
            )
            
        except Exception as e:
            logger.error(f"Error parsing generic CSV: {e}")
            return ParseResult(
                transactions=[],
                total_rows=0,
                parsed_rows=0,
                skipped_rows=0,
                errors=[f"Failed to parse file: {str(e)}"],
                warnings=[],
            )


class ExcelParser(BankCsvParser):
    """Parser for Excel (.xlsx, .xls) bank statement exports."""

    def detect(self, content: bytes) -> bool:
        """Detect if content is an Excel file."""
        # Excel files start with specific magic bytes
        # XLSX (ZIP-based): PK (50 4B)
        # XLS: D0 CF 11 E0
        if content[:2] == b'PK':  # XLSX
            return True
        if content[:4] == b'\xD0\xCF\x11\xE0':  # XLS
            return True
        return False

    def parse(self, content: bytes, filename: str = "") -> ParseResult:
        """Parse Excel content."""
        if not EXCEL_SUPPORT:
            return ParseResult(
                transactions=[],
                total_rows=0,
                parsed_rows=0,
                skipped_rows=0,
                errors=["Excel support not available. Please install openpyxl: pip install openpyxl"],
                warnings=[],
            )

        transactions = []
        errors = []
        warnings = []
        skipped_rows = 0
        
        try:
            # Load workbook from bytes
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            
            if sheet is None:
                return ParseResult(
                    transactions=[],
                    total_rows=0,
                    parsed_rows=0,
                    skipped_rows=0,
                    errors=["No active sheet found in Excel file"],
                    warnings=[],
                )

            # Get headers from first row
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value or "").strip().lower())
            
            # Map columns
            date_col = self._find_column_index(headers, ["date", "trans date", "transaction date", "posting date"])
            desc_col = self._find_column_index(headers, ["description", "desc", "memo", "narrative", "details", "payee"])
            amount_col = self._find_column_index(headers, ["amount", "amt", "value", "transaction amount"])
            credit_col = self._find_column_index(headers, ["credit", "deposit", "cr", "credits", "money in"])
            debit_col = self._find_column_index(headers, ["debit", "withdrawal", "dr", "debits", "money out"])
            balance_col = self._find_column_index(headers, ["balance", "running balance", "available balance"])
            type_col = self._find_column_index(headers, ["type", "transaction type", "dr/cr"])
            check_col = self._find_column_index(headers, ["check", "check number", "check #", "cheque"])
            
            if date_col is None:
                return ParseResult(
                    transactions=[],
                    total_rows=0,
                    parsed_rows=0,
                    skipped_rows=0,
                    errors=[f"Could not identify date column. Headers found: {headers}"],
                    warnings=[],
                )

            row_num = 0
            min_date = None
            max_date = None
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_num += 1
                
                try:
                    # Parse date
                    date_val = row[date_col] if date_col < len(row) else None
                    date = self._parse_excel_date(date_val)
                    
                    if not date:
                        skipped_rows += 1
                        continue
                    
                    if min_date is None or date < min_date:
                        min_date = date
                    if max_date is None or date > max_date:
                        max_date = date
                    
                    # Parse description
                    description = str(row[desc_col] if desc_col is not None and desc_col < len(row) else "Unknown").strip()
                    if not description:
                        description = "Unknown Transaction"
                    
                    # Parse amount and determine type
                    amount = None
                    txn_type = "debit"
                    
                    if amount_col is not None and amount_col < len(row):
                        amount = self._parse_excel_amount(row[amount_col])
                        if amount is not None:
                            txn_type = "credit" if amount > 0 else "debit"
                            amount = abs(amount)
                    elif credit_col is not None and debit_col is not None:
                        credit = self._parse_excel_amount(row[credit_col] if credit_col < len(row) else None)
                        debit = self._parse_excel_amount(row[debit_col] if debit_col < len(row) else None)
                        
                        if credit and credit > 0:
                            amount = credit
                            txn_type = "credit"
                        elif debit and debit > 0:
                            amount = debit
                            txn_type = "debit"
                    
                    if amount is None or amount == 0:
                        skipped_rows += 1
                        continue
                    
                    # Override type if column exists
                    if type_col is not None and type_col < len(row) and row[type_col]:
                        type_val = str(row[type_col]).lower()
                        if "credit" in type_val or "cr" in type_val or "deposit" in type_val:
                            txn_type = "credit"
                        elif "debit" in type_val or "dr" in type_val or "withdrawal" in type_val:
                            txn_type = "debit"
                    
                    # Parse balance
                    balance = None
                    if balance_col is not None and balance_col < len(row):
                        balance = self._parse_excel_amount(row[balance_col])
                    
                    # Parse check number
                    check_number = None
                    if check_col is not None and check_col < len(row) and row[check_col]:
                        check_number = str(row[check_col]).strip()
                    
                    # Build raw data dict
                    raw_data = {}
                    for i, h in enumerate(headers):
                        if i < len(row):
                            val = row[i]
                            if isinstance(val, datetime):
                                val = val.isoformat()
                            raw_data[h] = val
                    
                    txn = ParsedTransaction(
                        date=date,
                        description=description,
                        amount=amount,
                        type=txn_type,
                        balance=balance,
                        check_number=check_number,
                        raw_data=raw_data,
                        row_number=row_num + 1,  # +1 for header row
                    )
                    
                    transactions.append(txn)
                    
                except Exception as e:
                    errors.append(f"Row {row_num + 1}: {str(e)}")
                    skipped_rows += 1
            
            wb.close()
            
            return ParseResult(
                transactions=transactions,
                total_rows=row_num,
                parsed_rows=len(transactions),
                skipped_rows=skipped_rows,
                errors=errors,
                warnings=warnings,
                bank_name=self._detect_bank_from_filename(filename),
                statement_start=min_date,
                statement_end=max_date,
            )
            
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            return ParseResult(
                transactions=[],
                total_rows=0,
                parsed_rows=0,
                skipped_rows=0,
                errors=[f"Failed to parse Excel file: {str(e)}"],
                warnings=[],
            )

    def _find_column_index(self, headers: List[str], candidates: List[str]) -> Optional[int]:
        """Find column index matching any candidate."""
        for i, h in enumerate(headers):
            h_lower = h.lower().strip()
            for candidate in candidates:
                if candidate in h_lower or h_lower in candidate:
                    return i
        return None

    def _parse_excel_date(self, value: Any) -> Optional[datetime]:
        """Parse date from Excel cell value."""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value
        
        if isinstance(value, (int, float)):
            # Excel serial date
            try:
                from openpyxl.utils.datetime import from_excel
                return from_excel(value)
            except Exception:
                pass
        
        # Try parsing as string
        return self._parse_date(str(value))

    def _parse_excel_amount(self, value: Any) -> Optional[float]:
        """Parse amount from Excel cell value."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        return self._parse_amount(str(value))

    def _detect_bank_from_filename(self, filename: str) -> str:
        """Try to detect bank name from filename."""
        filename_lower = filename.lower()
        banks = ["chase", "bofa", "bank of america", "wells fargo", "citi", "capital one", "pnc", "usaa"]
        for bank in banks:
            if bank in filename_lower:
                return bank.title()
        return "Unknown"


class ZipParser(BankCsvParser):
    """Parser for ZIP files containing bank statements."""

    def detect(self, content: bytes) -> bool:
        """Detect if content is a ZIP file."""
        return content[:2] == b'PK' and zipfile.is_zipfile(io.BytesIO(content))

    def parse(self, content: bytes, filename: str = "") -> ParseResult:
        """Parse ZIP file containing bank statements."""
        all_transactions = []
        all_errors = []
        all_warnings = []
        total_rows = 0
        skipped_rows = 0
        min_date = None
        max_date = None
        
        try:
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                for name in zf.namelist():
                    # Skip directories and hidden files
                    if name.endswith('/') or name.startswith('__MACOSX') or name.startswith('.'):
                        continue
                    
                    ext = name.lower().split('.')[-1]
                    if ext not in ['csv', 'xlsx', 'xls']:
                        all_warnings.append(f"Skipped unsupported file: {name}")
                        continue
                    
                    # Read file content
                    file_content = zf.read(name)
                    
                    # Get appropriate parser
                    parser = get_parser_for_content(file_content)
                    result = parser.parse(file_content, name)
                    
                    # Merge results
                    all_transactions.extend(result.transactions)
                    all_errors.extend([f"{name}: {e}" for e in result.errors])
                    all_warnings.extend([f"{name}: {w}" for w in result.warnings])
                    total_rows += result.total_rows
                    skipped_rows += result.skipped_rows
                    
                    if result.statement_start:
                        if min_date is None or result.statement_start < min_date:
                            min_date = result.statement_start
                    if result.statement_end:
                        if max_date is None or result.statement_end > max_date:
                            max_date = result.statement_end
            
            return ParseResult(
                transactions=all_transactions,
                total_rows=total_rows,
                parsed_rows=len(all_transactions),
                skipped_rows=skipped_rows,
                errors=all_errors,
                warnings=all_warnings,
                bank_name="Multiple",
                statement_start=min_date,
                statement_end=max_date,
            )
            
        except Exception as e:
            logger.error(f"Error parsing ZIP file: {e}")
            return ParseResult(
                transactions=[],
                total_rows=0,
                parsed_rows=0,
                skipped_rows=0,
                errors=[f"Failed to parse ZIP file: {str(e)}"],
                warnings=[],
            )


def get_parser_for_content(content: bytes) -> BankCsvParser:
    """Get appropriate parser for the given content."""
    # Check for ZIP first (since XLSX is also ZIP-based)
    if content[:2] == b'PK':
        # Check if it's a valid ZIP with multiple files (not XLSX)
        try:
            with zipfile.ZipFile(io.BytesIO(content), 'r') as zf:
                names = zf.namelist()
                # XLSX files have specific internal structure
                if '[Content_Types].xml' in names or 'xl/workbook.xml' in names:
                    # This is an XLSX file
                    if EXCEL_SUPPORT:
                        return ExcelParser()
                else:
                    # This is a regular ZIP with multiple files
                    return ZipParser()
        except Exception:
            pass
    
    # Check for XLS (old Excel format)
    if content[:4] == b'\xD0\xCF\x11\xE0' and EXCEL_SUPPORT:
        return ExcelParser()
    
    # Try specific CSV parsers
    chase_parser = ChaseCsvParser()
    if chase_parser.detect(content):
        return chase_parser
    
    # Fall back to generic CSV
    return GenericCsvParser()
